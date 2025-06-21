from PyQt5.uic import loadUi
from PyQt5.QtGui import QColor, QIcon
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QTableWidgetItem
from mpl_toolkits.mplot3d import Axes3D
from scipy import interpolate
import numpy as np
import openpyxl
import xlrd
import math
import csv
import json
import chardet
import re
import sys
import os

BASE_DIR = os.path.dirname(__file__)
sys.path.insert(0, BASE_DIR)


class MainWindow(QMainWindow):

    def __init__(self):

        QMainWindow.__init__(self)

        loadUi(os.path.join(BASE_DIR, "ui/mainWindow.ui"), self)

        self.data = []
        self.deltaZ = []
        self.config = {}
        self.currentFolderPath = BASE_DIR
        self.currentFileType = self.sourceFileFormatSelect.currentText()
        self.initUi()
        self.initSiginal()

        # 载入默认配置文件
        self.defaultConfigFile = os.path.join(BASE_DIR, "default.tpl")
        if os.path.isfile(self.defaultConfigFile):
            self.loadConfig(self.defaultConfigFile)

    def initUi(self):
        self.updateBasePoints()
        self.sourceFileFormatSelect.setCurrentText(self.currentFileType)
        self.setWindowTitle("平整度分析程序 v1.0.3")
        self.setWindowIcon(QIcon(os.path.join(BASE_DIR, 'ui/icon.ico')))
        self.selectFileButton.setIcon(
            QIcon(os.path.join(BASE_DIR, 'ui/open.png')))
        self.btnLoadConfig.setIcon(
            QIcon(os.path.join(BASE_DIR, 'ui/configure.png')))
        self.btnSaveConfig.setIcon(
            QIcon(os.path.join(BASE_DIR, 'ui/save.png')))
        self.btnImportData.setIcon(
            QIcon(os.path.join(BASE_DIR, 'ui/search.png')))
        self.btnOutputData.setIcon(
            QIcon(os.path.join(BASE_DIR, 'ui/save.png')))
        self.btnShowGraph.setIcon(
            QIcon(os.path.join(BASE_DIR, 'ui/graph.png')))

    def initSiginal(self):
        self.selectFileButton.clicked.connect(self.selectFile)
        self.btnImportData.clicked.connect(self.importData)
        self.sourceFileFormatSelect.currentIndexChanged[int].connect(
            self.updateFileType)
        self.btnShowGraph.clicked.connect(self.nextTab)
        self.groupSize.valueChanged[int].connect(self.updateBasePoints)
        self.btnSaveConfig.clicked.connect(self.saveConfigFromDialog)
        self.btnLoadConfig.clicked.connect(self.loadConfigFromDialog)
        self.btnOutputData.clicked.connect(self.outputData)

        self.tabWidget.currentChanged[int].connect(self.tabChanged)
        self.flatnessTable.currentItemChanged.connect(self.changeDataGroup)
        self.localFlatnessTable.currentItemChanged.connect(self.showGraph1)
        self.currentIndexSelect.currentIndexChanged[int].connect(
            self.flatnessTable.selectRow)
        self.currentIndexSelect.currentIndexChanged[int].connect(
            self.currentIndexSelect1.setCurrentIndex)
        self.currentIndexSelect1.currentIndexChanged[int].connect(
            self.currentIndexSelect.setCurrentIndex)
        self.colorSelect.currentIndexChanged[int].connect(
            self.colorSelect1.setCurrentIndex)
        self.colorSelect1.currentIndexChanged[int].connect(
            self.colorSelect.setCurrentIndex)

        self.currentIndexSelect.currentIndexChanged[int].connect(
            self.showGraph)
        self.currentIndexSelect1.currentIndexChanged[int].connect(
            self.showGraph1)
        self.currentIndexSelect1.currentIndexChanged[int].connect(
            self.showLocalFlatness)
        self.colorSelect.currentIndexChanged[int].connect(self.showGraph)
        self.functionSelect.currentIndexChanged[int].connect(self.showGraph)
        self.colorSelect1.currentIndexChanged[int].connect(self.showGraph1)

    def closeEvent(self, event):
        if not os.path.isfile(self.defaultConfigFile):
            self.saveConfig(self.defaultConfigFile)
        event.accept()

    def formatList(self, points):
        startPoint = None
        endPoint = None
        result = []
        for curPoint in points:
            if startPoint is None:
                startPoint = curPoint
                endPoint = curPoint
            elif curPoint-endPoint == 1:
                endPoint = curPoint
            else:
                if endPoint is None or startPoint == endPoint:
                    result.append(str(startPoint))
                else:
                    result.append('{}-{}'.format(startPoint, endPoint))
                startPoint = curPoint
                endPoint = curPoint
        if startPoint is None:
            return ''
        elif startPoint == endPoint:
            result.append(str(startPoint))
        else:
            result.append('{}-{}'.format(startPoint, endPoint))
        return ','.join(result)


    def updateFileType(self, index):
        if index == 0:
            self.hasHeadlineButton.setEnabled(False)
            self.noHeadlineButton.setEnabled(False)
        else:
            self.hasHeadlineButton.setEnabled(True)
            self.noHeadlineButton.setEnabled(True)

    def tabChanged(self, id):
        if id == 1:
            self.showFlatness()
        elif id == 2:
            self.showGraph()
        elif id == 3:
            self.showLocalFlatness()
            self.showGraph1()

    def saveConfigFromDialog(self):
        filePath, _ = QFileDialog.getSaveFileName(self,
                                                  caption="保存参数配置",
                                                  filter="参数模板 (*.tpl)")
        if filePath:
            self.saveConfig(filePath)

    def saveConfig(self, filePath):
        config = {
            'SourceFileFormat': self.sourceFileFormatSelect.currentIndex(),
            'SourceFileHeader': self.hasHeadlineButton.isChecked(),
            'DataGroupSize': self.groupSize.value(),
            'AlignPoints': self.basePointsSelect.selectedIndex(),
            'NormalizeZ': self.normalizeSelect.currentIndex(),
            'FlatnessTolerance': self.flatnessTolerance.value(),
            'RbfFunction': self.functionSelect.currentIndex(),
            'GraphTheme': self.colorSelect.currentIndex(),
            'LocalFlatnessTolerance': self.localFlatnessTolerance.value(),
            'LocalFlatnessLimit': self.localFlatnessLimit.value(),
            'CentralZoneLimit': self.centralZoneLimit.value()
        }
        try:
            with open(filePath, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, sort_keys=False)
            return True
        except Exception as e:
            QMessageBox.critical(
                self, '错误', '无法保存参数配置文件！', QMessageBox.Ok)
            return False

    def loadConfigFromDialog(self):
        filePath, _ = QFileDialog.getOpenFileName(
            parent=self, caption="选择参数配置文件", filter="参数模板 (*.tpl)")
        if os.path.isfile(filePath):
            if self.loadConfig(filePath):
                QMessageBox.information(
                    self, '已载入', '已从模板文件中载入设置参数：\n'+filePath)

    def loadConfig(self, filePath):
        try:
            with open(filePath, mode='r', encoding='utf-8') as f:
                config = json.load(f)
                self.sourceFileFormatSelect.setCurrentIndex(
                    config.get('SourceFileFormat', 0))
                self.updateFileType(self.sourceFileFormatSelect.currentIndex())
                if config.get('SourceFileHeader', True):
                    self.hasHeadlineButton.setChecked(True)
                else:
                    self.noHeadlineButton.setChecked(True)
                self.groupSize.setValue(config.get('DataGroupSize', 4))
                alignPoints = config.get('AlignPoints', [])
                for i in range(1, self.basePointsSelect.row_num):
                    if i in alignPoints:
                        self.basePointsSelect.qCheckBox[i].setChecked(True)
                    else:
                        self.basePointsSelect.qCheckBox[i].setChecked(False)
                self.normalizeSelect.setCurrentIndex(
                    config.get('NormalizeZ', 0))
                self.flatnessTolerance.setValue(
                    config.get('FlatnessTolerance', 0.1))
                self.localFlatnessTolerance.setValue(
                    config.get('LocalFlatnessTolerance', 0.0504))
                self.localFlatnessLimit.setValue(
                    config.get('LocalFlatnessLimit', 25.4))
                self.functionSelect.setCurrentIndex(
                    config.get('RbfFunction', 0))
                self.colorSelect.setCurrentIndex(
                    config.get('GraphTheme', 0))
                self.centralZoneLimit.setValue(
                    config.get('CentralZoneLimit', 50))
            return True
        except Exception as e:
            QMessageBox.critical(
                self, '错误', '无法导入参数配置文件：\n'+str(e), QMessageBox.Ok)
            return False

    def updateBasePoints(self):
        # 更新理想平面参考点的复合下拉选择框
        items = ['量测点 '+str(i+1) for i in range(self.groupSize.value())]
        self.basePointsSelect.loadItems(items)
        self.basePointsSelect.selectAll()

    def changeDataGroup(self):
        # 点击平整度汇整表切换当前数据组别
        index = self.flatnessTable.currentItem().row()
        self.currentIndexSelect.setCurrentIndex(index)
        self.showFlatness()

    def dataAnalyze(self, rawdata):
        # 对数据进行分组计算平整度
        size = self.groupSize.value()
        if size > 0:
            if len(rawdata) % size == 0:
                data = [rawdata[i:i+size]
                        for i in range(0, len(rawdata), size)]
            else:
                QMessageBox.critical(
                    self, '错误', '导入的数据无法按要求进行分组，请确认源数据是否正确！', QMessageBox.Ok)
                return False
        else:
            data = [rawdata]

        # 保存当前设置参数
        self.config = {
            'groupSize': self.groupSize.value(),
            'refPoints': self.formatList(self.basePointsSelect.selectedIndex()) or '无',
            'normalizeZ': self.normalizeSelect.currentText(),
            'flatnessTolerance': self.flatnessTolerance.value(),
            'localFlatnessTolerance': self.localFlatnessTolerance.value(),
            'localFlatnessLimit': self.localFlatnessLimit.value(),
            'centralZoneLimit': self.centralZoneLimit.value()/100
        }
        if len(self.basePointsSelect.selectedIndex()) >= 3:
            # 当选择了三个或以上理想平面参考点时，计算理想平面方程系数
            for group in data:
                matrixA = []
                matrixB = []
                for i in self.basePointsSelect.selectedIndex():
                    point = group[i-1]
                    matrixA.append([point[0], point[1], 1])
                    matrixB.append([point[2]])
                matrixA = np.array(matrixA)
                matrixB = np.array(matrixB)
                matrixCoeff = np.dot(np.dot(np.linalg.inv(
                    np.dot(matrixA.T, matrixA)), matrixA.T), matrixB)
                coeffA = -1 * matrixCoeff[0][0]
                coeffB = -1 * matrixCoeff[1][0]
                coeffC = 1
                coeffD = -1 * matrixCoeff[2][0]
                constant = math.sqrt(coeffA*coeffA+coeffB*coeffB+coeffC*coeffC)

                minX = None
                maxX = None
                minY = None
                maxY = None
                for point in group:
                    point.append(
                        (coeffA*point[0]+coeffB*point[1]+coeffC*point[2]+coeffD)/constant)
                    if minX is None or point[0] < minX:
                        minX = point[0]
                    if maxX is None or point[0] > maxX:
                        maxX = point[0]
                    if minY is None or point[1] < minY:
                        minY = point[1]
                    if maxY is None or point[1] > maxY:
                        maxY = point[1]

                group[0].append({
                    'A': coeffA,
                    'B': coeffB,
                    'C': coeffC,
                    'D': coeffD,
                    'minX': minX,
                    'maxX': maxX,
                    'minY': minY,
                    'maxY': maxY
                })
        else:
            # 如果不计算理想参考平面，直接使用量测的Z值进行计算，平面方程 Z=0
            for group in data:
                minX = None
                maxX = None
                minY = None
                maxY = None
                for point in group:
                    point.append(point[2])  # Z'
                    if minX is None or point[0] < minX:
                        minX = point[0]
                    if maxX is None or point[0] > maxX:
                        maxX = point[0]
                    if minY is None or point[1] < minY:
                        minY = point[1]
                    if maxY is None or point[1] > maxY:
                        maxY = point[1]

                group[0].append({
                    'A': 0,
                    'B': 0,
                    'C': 1,
                    'D': 0,
                    'minX': minX,
                    'maxX': maxX,
                    'minY': minY,
                    'maxY': maxY
                })

        # 正规化处理 Z 值并更新平整度结果
        normalizeType = self.normalizeSelect.currentIndex()
        row = 0

        for group in data:
            # 数据Z'统计值
            min = None
            max = None
            sum = 0

            # 计算中心形貌统计值
            centralMinZ = None
            centralMaxZ = None
            marginalSum = 0
            marginalCount = 0

            # 迭代计算最小、最大、加总、平均值
            for point in group:
                # 计算总体Z'的统计值
                if min is None or point[3] < min:
                    min = point[3]
                if max is None or point[3] > max:
                    max = point[3]
                sum += point[3]

                # 计算中心区域Z'统计值
                minX = group[0][4]['minX']
                maxX = group[0][4]['maxX']
                minY = group[0][4]['minY']
                maxY = group[0][4]['maxY']
                rangeX = maxX-minX
                rangeY = maxY-minY

                if abs(2*(point[0]-minX)/rangeX - 1) < self.config['centralZoneLimit'] and abs(2*(point[1]-minY)/rangeY - 1) < self.config['centralZoneLimit']:
                    # 当前量测点为板中心位置时
                    if centralMinZ is None or point[3] < centralMinZ:
                        centralMinZ = point[3]
                    if centralMaxZ is None or point[3] > centralMaxZ:
                        centralMaxZ = point[3]
                else:
                    # 当前量测点为板边位置时
                    marginalSum += point[3]
                    marginalCount += 1

            # 计算总体平均值
            avg = sum/len(group)
            marginalAvg = marginalSum/marginalCount
            group[0][4]['shape'] = 'unknow'
            if centralMinZ is not None:
                if centralMinZ > marginalAvg and centralMaxZ > marginalAvg:
                    # 中心凸起
                    group[0][4]['shape'] = 'convex'
                elif centralMinZ < marginalAvg and centralMaxZ < marginalAvg:
                    # 中心下凹
                    group[0][4]['shape'] = 'concave'
                else:
                    # 凹凸不平
                    group[0][4]['shape'] = 'wavy convex'
            if normalizeType == 0:  # 不处理
                offset = 0
            elif normalizeType == 1:  # 正值化
                offset = min
            elif normalizeType == 2:  # 负值化
                offset = max
            elif normalizeType == 3:  # 均值中心
                offset = avg
            elif normalizeType == 4:  # 极差中心
                offset = (min+max)/2
            for point in group:
                point[3] = point[3]-offset

            group[0][4]['min'] = min-offset
            group[0][4]['max'] = max-offset
            group[0][4]['flatness'] = max-min
        self.data = data

        # 更新平整度数据分析结果汇总表
        row = 0
        self.flatnessTable.setRowCount(len(data))
        for group in data:
            flatness = group[0][4]['flatness']
            item0 = QTableWidgetItem('{:.4f}'.format(group[0][4]['min']))
            item1 = QTableWidgetItem('{:.4f}'.format(group[0][4]['max']))
            item2 = QTableWidgetItem('{:.4f}'.format(flatness))
            item3 = QTableWidgetItem(group[0][4]['shape'])
            if flatness <= self.flatnessTolerance.value():
                item4 = QTableWidgetItem('Acc')
                item4.setBackground(QColor(100, 200, 100))
                group[0][4]['judge'] = 'Acc'
            else:
                item4 = QTableWidgetItem('Rej')
                item4.setBackground(QColor(200, 50, 50))
                group[0][4]['judge'] = 'Rej'

            item0.setTextAlignment(Qt.AlignCenter)
            item1.setTextAlignment(Qt.AlignCenter)
            item2.setTextAlignment(Qt.AlignCenter)
            item3.setTextAlignment(Qt.AlignCenter)
            item4.setTextAlignment(Qt.AlignCenter)

            self.flatnessTable.setItem(row, 0, item0)
            self.flatnessTable.setItem(row, 1, item1)
            self.flatnessTable.setItem(row, 2, item2)
            self.flatnessTable.setItem(row, 3, item3)
            self.flatnessTable.setItem(row, 4, item4)
            row += 1

        # 计算局部平整度
        localFlatnessLimit = self.localFlatnessLimit.value()
        self.deltaZ = []
        for group in data:
            count = len(group)
            groupDeltaZ = []
            for i in range(count):
                for j in range(i+1, count):
                    p1 = group[i]
                    p2 = group[j]
                    distance = math.sqrt(
                        math.pow(p1[0]-p2[0], 2)+math.pow(p1[1]-p2[1], 2))
                    if distance <= localFlatnessLimit:
                        groupDeltaZ.append([i, j, distance, abs(p2[3]-p1[3])])
            self.deltaZ.append(groupDeltaZ)

        # 更新当前选择的数据组别
        self.currentIndexSelect.clear()
        self.currentIndexSelect1.clear()
        groups = ["第 {} 组".format(i+1) for i in range(len(data))]
        self.currentIndexSelect.addItems(groups)
        self.currentIndexSelect1.addItems(groups)
        self.currentIndexSelect.setCurrentIndex(0)

    def load_csv(self, file_path):
        # 导入CSV格式的数据文件
        data = []
        with open(file_path) as f:
            csv_reader = csv.reader(f)
            if self.hasHeadlineButton.isChecked():
                next(csv_reader)
            for row in csv_reader:
                if len(row) >= 3:
                    data.append([float(row[0]), float(row[1]), float(row[2])])
        return data

    def load_xlsx(self, file_path):
        # 导入 .xlsx Excel工作簿
        data = []
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        row = 2 if self.hasHeadlineButton.isChecked() else 1
        while True:
            x = ws.cell(row=row, column=1).value
            y = ws.cell(row=row, column=2).value
            z = ws.cell(row=row, column=3).value
            if x is None or y is None or z is None:
                break
            data.append([float(x), float(y), float(z)])
            row += 1
        return data

    def load_xls(self, file_path):
        # 导入旧的 .xls Excel工作簿
        data = []
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheets()[0]
        row = 1 if self.hasHeadlineButton.isChecked() else 0
        while True:
            x = ws.cell(row, 0)
            y = ws.cell(row, 1)
            z = ws.cell(row, 2)
            if x.ctype != 2 or y.ctype != 2 or z.ctype != 2:
                break
            data.append([float(x.value), float(y.value), float(z.value)])
            row += 1
            if row >= ws.nrows:
                break
        return data

    def load_txt(self, file_path):
        # 导入三次元测量数据 .txt 文件
        data = []
        with open(file_path, mode='rb') as f:
            encoding = chardet.detect(f.read())['encoding']
        pattern = re.compile(
            r'.*X 坐标\s+(-?\d+\.?\d*).*Y 坐标\s+(-?\d+\.?\d*).*Z 坐标\s+(-?\d+\.?\d*).*')
        with open(file_path, mode='r', encoding=encoding, errors='ignore') as f:
            for line in f:
                result = pattern.match(line)
                if result:
                    x, y, z = result.groups()
                    data.append(
                        [float(x), float(y), float(z)])
        return data

    def importData(self):
        # 导入数据文件
        filePath = self.selectedFilePath.text()
        if not filePath:
            QMessageBox.critical(
                self, '错误', '请选择需要导入的数据文件！', QMessageBox.Ok)
            return
        try:
            refPoints = self.basePointsSelect.selectedIndex()
            if 0 < len(refPoints) < 3:
                QMessageBox.critical(
                    self, '错误', '请选择平面基准点，参数点数少于3个时无法计算理想平面！', QMessageBox.Ok)
                return

            data = None
            if self.currentFileType == 'Excel工作簿 (*.xlsx)':
                data = self.load_xlsx(filePath)
            elif self.currentFileType == '旧Excel工作簿 (*.xls)':
                data = self.load_xls(filePath)
            elif self.currentFileType == 'CSV文本文件 (*.csv)':
                data = self.load_csv(filePath)
            elif self.currentFileType == '三次元测量数据 (*.txt)':
                data = self.load_txt(filePath)
            else:
                QMessageBox.critical(
                    self, '错误', '没有合适的处理器以导入当前的数据文件!', QMessageBox.Ok)
                return
            if not data:
                raise ValueError('导入的数据文件为空！')
            self.dataAnalyze(data)
            self.tabWidget.setCurrentIndex(1)
        except Exception as e:
            QMessageBox.critical(
                self, '错误', '读取量测数据时出现错误，请确认选择的数据文件格式是否正确！\n'+str(e), QMessageBox.Ok)

    def outputData(self):
        # 保存平整度计算结果
        if len(self.data) == 0:
            return

        filePath, _ = QFileDialog.getSaveFileName(
            parent=self,
            caption="保存平整度计算结果",
            directory=self.currentFolderPath,
            filter='Excel工作簿 (*.xlsx)'
        )
        if not filePath:
            return
        try:
            wb = openpyxl.Workbook()
            ws1 = wb.active

            highlight = openpyxl.styles.Font(color='FF0000')
            bold = openpyxl.styles.Font(bold=True)
            alignleft = openpyxl.styles.Alignment(
                horizontal='left', vertical='center')
            aligncenter = openpyxl.styles.Alignment(
                horizontal='center', vertical='center')

            ws1.title = "数据汇总"
            ws2 = wb.create_sheet(title="数据明细")
            ws3 = wb.create_sheet(title="局部平整度")

            ws1.column_dimensions['A'].width = 15
            ws1.cell(row=1, column=1, value='分组数')
            ws1.cell(row=2, column=1, value='每组量测点数')
            ws1.cell(row=3, column=1, value='平面参考点')
            ws1.cell(row=4, column=1, value="Z'标准化")
            ws1.cell(row=5, column=1, value='整板平整度公差')
            ws1.cell(row=6, column=1, value='局部平整度公差')
            ws1.cell(row=7, column=1, value='中心区域范围')
            ws1.cell(row=1, column=2, value=len(self.data))
            ws1.cell(row=1, column=2).alignment = alignleft
            ws1.cell(row=2, column=2, value=self.config['groupSize'])
            ws1.cell(row=2, column=2).alignment = alignleft
            ws1.cell(row=3, column=2, value=self.config['refPoints'])
            ws1.cell(row=4, column=2, value=self.config['normalizeZ'])
            ws1.cell(row=5, column=2, value=self.config['flatnessTolerance'])
            ws1.cell(row=5, column=2).alignment = alignleft
            ws1.cell(row=6, column=2, value='{} / {}'.format(
                self.config['localFlatnessTolerance'], self.config['localFlatnessLimit']))
            ws1.cell(row=7, column=2, value='{}%'.format(
                int(self.config['centralZoneLimit']*100)))

            ws1.cell(row=9, column=1, value='分组编号')
            ws1.cell(row=9, column=2, value="最小Z'")
            ws1.cell(row=9, column=3, value="最大Z'")
            ws1.cell(row=9, column=4, value='平整度')
            ws1.cell(row=9, column=5, value='中心形貌')
            ws1.cell(row=9, column=6, value='判定')
            ws1.cell(row=9, column=6).alignment = aligncenter
            ws1.cell(row=9, column=7, value='参考平面方程')

            ws2.cell(row=1, column=1, value='分组编号')
            ws2.cell(row=1, column=2, value='量测点')
            ws2.cell(row=1, column=3, value='X')
            ws2.cell(row=1, column=4, value='Y')
            ws2.cell(row=1, column=5, value='Z')
            ws2.cell(row=1, column=6, value="Z'")

            ws3.cell(row=1, column=1, value='分组编号')
            ws3.cell(row=1, column=2, value='量测点1')
            ws3.cell(row=1, column=3, value='量测点2')
            ws3.cell(row=1, column=4, value='平面距离')
            ws3.cell(row=1, column=5, value='平整度')
            ws3.cell(row=1, column=6, value='判定')
            ws3.cell(row=1, column=6).alignment = aligncenter

            row1 = 2
            row2 = 2
            for i in range(len(self.data)):
                group = self.data[i]
                deltaZ = self.deltaZ[i]
                info = group[0][4]
                ws1.cell(row=i+10, column=1, value=i+1)
                ws1.cell(row=i+10, column=2, value=info['min'])
                ws1.cell(row=i+10, column=3, value=info['max'])
                ws1.cell(row=i+10, column=4, value=info['flatness'])
                ws1.cell(row=i+10, column=5, value=info['shape'])
                ws1.cell(row=i+10, column=6, value=info['judge'])
                ws1.cell(row=i+10, column=6).alignment = aligncenter
                if info['judge'] == 'Rej':
                    ws1.cell(row=i+10, column=6).font = highlight
                ws1.cell(row=i+10, column=7, value='Z = {:.8e}X {:+.8e}Y {:+.8e}'.format(
                    -info['A'],
                    -info['B'],
                    -info['D'],
                ))
                for j in range(len(group)):
                    point = group[j]
                    ws2.cell(row=row1, column=1, value=i+1)
                    ws2.cell(row=row1, column=2, value=j+1)
                    ws2.cell(row=row1, column=3, value=point[0])
                    ws2.cell(row=row1, column=4, value=point[1])
                    ws2.cell(row=row1, column=5, value=point[2])
                    ws2.cell(row=row1, column=6, value=point[3])

                    minX = group[0][4]['minX']
                    maxX = group[0][4]['maxX']
                    minY = group[0][4]['minY']
                    maxY = group[0][4]['maxY']
                    if abs(2*(point[0]-minX)/(maxX-minX) - 1) < self.config['centralZoneLimit'] and abs(2*(point[1]-minY)/(maxY-minY) - 1) < self.config['centralZoneLimit']:
                        # 如果为中心区域量测点，将编号加粗
                        ws2.cell(row=row1, column=2).font = bold
                        ws2.cell(row=row1, column=3).font = bold
                        ws2.cell(row=row1, column=4).font = bold
                        ws2.cell(row=row1, column=5).font = bold
                        ws2.cell(row=row1, column=6).font = bold

                    row1 += 1
                for v in deltaZ:
                    ws3.cell(row=row2, column=1, value=i+1)
                    ws3.cell(row=row2, column=2, value=v[0]+1)
                    ws3.cell(row=row2, column=3, value=v[1]+1)
                    ws3.cell(row=row2, column=4, value=v[2])
                    ws3.cell(row=row2, column=5, value=v[3])

                    if v[3] <= self.config['localFlatnessTolerance']:
                        ws3.cell(row=row2, column=6, value='Acc')
                    else:
                        ws3.cell(row=row2, column=6, value='Rej')
                        ws3.cell(row=row2, column=6).font = highlight
                    ws3.cell(row=row2, column=6).alignment = aligncenter
                    row2 += 1
            wb.save(filename=filePath)
            QMessageBox.information(self, '已完成', '数据已保存至以下文件：\n'+filePath)
        except Exception as e:
            QMessageBox.critical(
                self, '错误', '无法保存当前数据至指定文件：\n'+filePath, QMessageBox.Ok)

    def selectFile(self):
        # 选择要导入的数据文件
        filter = self.sourceFileFormatSelect.currentText()
        filters = []
        for i in range(self.sourceFileFormatSelect.count()):
            filters.append(self.sourceFileFormatSelect.itemText(i))
        filePath, fileType = QFileDialog.getOpenFileName(
            parent=self,
            caption="选择平整度量测数据",
            directory=self.currentFolderPath,
            filter=';;'.join(filters),
            initialFilter=filter
        )
        if os.path.isfile(filePath):
            self.selectedFilePath.setText(filePath)
            self.currentFolderPath = os.path.split(filePath)[0]
            self.currentFileType = fileType
            self.sourceFileFormatSelect.setCurrentText(fileType)

    def nextTab(self):
        # 切换到下一个标签页
        index = self.tabWidget.currentIndex()
        index += 1
        if index > self.tabWidget.count()-1:
            index = 0
        self.tabWidget.setCurrentIndex(index)

    def showFlatness(self):
        # 显示量测点Z'数据明细
        index = self.currentIndexSelect.currentIndex()
        if index < 0 or len(self.data) == 0 or self.tabWidget.currentIndex() != 1:
            return
        data = self.data[index]
        row = 0
        self.dataTable.setRowCount(len(data))
        for point in data:
            item0 = QTableWidgetItem('{:.4f}'.format(point[0]))
            item1 = QTableWidgetItem('{:.4f}'.format(point[1]))
            item2 = QTableWidgetItem('{:.4f}'.format(point[2]))
            item3 = QTableWidgetItem('{:.4f}'.format(point[3]))
            item0.setTextAlignment(Qt.AlignCenter)
            item1.setTextAlignment(Qt.AlignCenter)
            item2.setTextAlignment(Qt.AlignCenter)
            item3.setTextAlignment(Qt.AlignCenter)

            color = (point[3]-data[0][4]['min'])/data[0][4]['flatness']*2-1
            if color <= 0:
                color = int(255+170*color)
                item3.setBackground(QColor(color, color, 255))
            else:
                color = int(255-170*color)
                item3.setBackground(QColor(255, color, color))

            self.dataTable.setItem(row, 0, item0)
            self.dataTable.setItem(row, 1, item1)
            self.dataTable.setItem(row, 2, item2)
            self.dataTable.setItem(row, 3, item3)
            row += 1

    def showLocalFlatness(self):
        # 显示两点间的局部平整度数据表
        index = self.currentIndexSelect.currentIndex()
        if index < 0 or index >= len(self.deltaZ):
            return
        deltaZ = self.deltaZ[index]
        self.localFlatnessTable.setRowCount(len(deltaZ))
        row = 0
        for r in deltaZ:
            item0 = QTableWidgetItem('{} - {}'.format(r[0]+1, r[1]+1))
            item1 = QTableWidgetItem('{:.4f}'.format(r[2]))
            item2 = QTableWidgetItem('{:.4f}'.format(r[3]))
            if abs(r[3]) <= self.config['localFlatnessTolerance']:
                item3 = QTableWidgetItem('Acc')
                item3.setBackground(QColor(100, 200, 100))
            else:
                item3 = QTableWidgetItem('Rej')
                item3.setBackground(QColor(200, 50, 50))

            item0.setTextAlignment(Qt.AlignCenter)
            item1.setTextAlignment(Qt.AlignCenter)
            item2.setTextAlignment(Qt.AlignCenter)
            item3.setTextAlignment(Qt.AlignCenter)
            self.localFlatnessTable.setItem(row, 0, item0)
            self.localFlatnessTable.setItem(row, 1, item1)
            self.localFlatnessTable.setItem(row, 2, item2)
            self.localFlatnessTable.setItem(row, 3, item3)
            row += 1

    def getAxesLimit(self, serialx, serialy):
        # 根据X、Y坐标数据计算图表坐标轴显示范围
        xmin = np.min(serialx)
        xmax = np.max(serialx)
        xavg = (xmin+xmax)/2
        ymin = np.min(serialy)
        ymax = np.max(serialy)
        yavg = (ymin+ymax)/2
        xrange = xmax-xmin
        yrange = ymax-ymin
        if xrange > yrange:
            return xmin, xmax, yavg-yrange*xrange/yrange/2, yavg+yrange*xrange/yrange/2
        else:
            return xavg-xrange*yrange/xrange/2, xavg+xrange*yrange/xrange/2, ymin, ymax

    def showGraph(self):
        # 显示三维曲面图
        if self.tabWidget.currentIndex() != 2:
            return
        self.MplWidget.figure.clear()
        self.MplWidget.axes = Axes3D(
            self.MplWidget.figure, auto_add_to_figure=False)
        self.MplWidget.figure.add_axes(self.MplWidget.axes)
        self.MplWidget.axes.set_xlabel('X')
        self.MplWidget.axes.set_ylabel('Y')

        if len(self.data) == 0:
            return
        data = self.data[self.currentIndexSelect.currentIndex()]
        x = [v[0] for v in data]
        y = [v[1] for v in data]
        z = [v[3] for v in data]

        func_name = self.functionSelect.currentText()
        color_map = self.colorSelect.currentText()

        func = interpolate.Rbf(x, y, z, function=func_name)
        xnew, ynew = np.mgrid[np.min(x):np.max(x):50j, np.min(y):np.max(y):50j]
        znew = func(xnew, ynew)
        newz = func(x, y)
        self.MplWidget.axes.scatter(x, y, newz+0.0001, c='r', marker='o')

        surf = self.MplWidget.axes.plot_surface(
            xnew, ynew, znew, cmap=color_map)
        self.MplWidget.colorbar = self.MplWidget.figure.colorbar(
            surf, shrink=0.6, aspect=10)

        # 设定X、Y坐标轴范围
        minX, maxX, minY, maxY = self.getAxesLimit(x, y)
        self.MplWidget.axes.set_xlim(minX, maxX)
        self.MplWidget.axes.set_ylim(minY, maxY)

        self.MplWidget.canvas.draw()

    def showGraph1(self):
        # 显示轮廓图
        if self.tabWidget.currentIndex() != 3:
            return
        groupIndex = self.currentIndexSelect.currentIndex()
        self.MplWidget1.figure.clear()
        self.MplWidget1.axes = self.MplWidget1.figure.add_subplot()
        self.MplWidget1.axes.set_xlabel('X')
        self.MplWidget1.axes.set_ylabel('Y')
        if len(self.data) == 0:
            return
        data = self.data[groupIndex]
        x = [v[0] for v in data]
        y = [v[1] for v in data]
        z = [v[3] for v in data]

        func_name = self.functionSelect.currentText()
        color_map = self.colorSelect.currentText()

        func = interpolate.Rbf(x, y, z, function=func_name)
        xnew, ynew = np.mgrid[np.min(x):np.max(x):50j, np.min(y):np.max(y):50j]
        znew = func(xnew, ynew)

        contour = self.MplWidget1.axes.contourf(
            xnew, ynew, znew, cmap=color_map)

        # 设定X、Y坐标轴范围
        minX, maxX, minY, maxY = self.getAxesLimit(x, y)
        self.MplWidget1.axes.set_xlim(minX, maxX)
        self.MplWidget1.axes.set_ylim(minY, maxY)

        self.MplWidget1.axes.scatter(x, y,  c='r', marker='o')
        currentRow = self.localFlatnessTable.currentItem()
        if currentRow is not None:
            index = currentRow.row()
            r = self.deltaZ[groupIndex][index]
            p1 = data[r[0]]
            p2 = data[r[1]]

            self.MplWidget1.axes.quiver(
                p1[0], p1[1], p2[0]-p1[0], p2[1]-p1[1],  angles='xy', scale=1, scale_units='xy')

        self.MplWidget1.canvas.draw()


app = QApplication([])
window = MainWindow()
window.show()
app.exec_()
